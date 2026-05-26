"""Exporta a Excel el listado 'En curso' de Peritaciones Diversos desde ePAC.

Flujo:
1. Obtiene URL y credenciales de ePAC desde la BD (con fallback a variables de entorno).
2. Lanza Playwright, hace login y navega a Peritaciones Diversos.
3. Selecciona la pestaña 'En curso' y envía el formulario vacío.
4. Extrae el listado paginado y guarda los registros en un fichero Excel.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from pathlib import Path
from typing import Optional

# ─── Añadir raíz del repo de módulos compartidos al path ─────────────────────
for _candidate in ("contacto_whatsapp", "preliminares-upload"):
    _EPAC_ROOT = Path(__file__).resolve().parent.parent / _candidate
    if _EPAC_ROOT.is_dir():
        if str(_EPAC_ROOT) not in sys.path:
            sys.path.insert(0, str(_EPAC_ROOT))
        break
# ─────────────────────────────────────────────────────────────────────────────

try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(dotenv_path=Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

try:
    import mysql.connector as _mysql_connector  # type: ignore
    _mysql_error: Optional[ImportError] = None
except ImportError:
    _mysql_connector = None  # type: ignore
    _mysql_error = ImportError("mysql-connector-python no está instalado.")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from playwright.sync_api import Page, expect

from browser import launch_browser
from config import load_config
from epac.pages.login_page import LoginPage
from epac.pages.navigation_page import NavigationPage


COLUMNAS: list[str] = [
    "Stro./Encargo",
    "A.R.",
    "Tipo",
    "F.Encargo",
    "Prox.Visita",
    "Referencia",
    "Población",
    "Provincia",
    "C.Postal",
    "I.",
]

OUTPUT_DIR: Path = Path.home() / "preliminares-extraction"

_IFRAME_SELECTOR: str = "iframe[name='appArea']"

_ANCHOS_COLUMNA: dict[str, int] = {
    "Stro./Encargo": 18,
    "A.R.": 8,
    "Tipo": 10,
    "F.Encargo": 14,
    "Prox.Visita": 14,
    "Referencia": 22,
    "Población": 20,
    "Provincia": 15,
    "C.Postal": 10,
    "I.": 8,
}


def _cargar_env(ruta_env: Optional[str]) -> None:
    """Carga variables de entorno desde un fichero .env alternativo.

    Args:
        ruta_env: Ruta al fichero .env. Si es None, no hace nada.

    Raises:
        FileNotFoundError: Si la ruta indicada no existe.
    """
    if not ruta_env:
        return
    ruta = Path(ruta_env)
    if not ruta.exists():
        raise FileNotFoundError(f"No existe el fichero --env: {ruta}")
    for linea in ruta.read_text(encoding="utf-8").splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#") or "=" not in linea:
            continue
        clave, valor = linea.split("=", 1)
        os.environ.setdefault(
            clave.strip(), valor.strip().strip('"').strip("'")
        )


def _construir_params_bd() -> dict:
    """Construye los parámetros de conexión a la BD incluyendo SSL si está configurado.

    Returns:
        Diccionario con los parámetros para mysql.connector.connect.
    """
    params: dict = {
        "host": os.environ.get("DB_HOST"),
        "port": int(os.environ.get("DB_PORT", 3306)),
        "user": os.environ.get("DB_USER"),
        "password": os.environ.get("DB_PASS"),
        "database": os.environ.get("DB_NAME"),
        "connection_timeout": 10,
        "autocommit": True,
        "use_pure": True,
        "charset": "latin1",
    }
    ssl_ca = os.environ.get("DB_SSL_CA")
    ssl_cert = os.environ.get("DB_SSL_CERT")
    ssl_key = os.environ.get("DB_SSL_KEY")
    if ssl_ca or ssl_cert or ssl_key:
        params.update(
            {
                "ssl_ca": ssl_ca,
                "ssl_cert": ssl_cert,
                "ssl_key": ssl_key,
                "ssl_disabled": False,
                "ssl_verify_cert": False,
                "ssl_verify_identity": False,
            }
        )
    return params


def obtener_credenciales_epac() -> tuple[str, str, str]:
    """Obtiene URL y credenciales de ePAC desde la BD, con fallback al entorno.

    Consulta softline_aseguradoras_claves_web filtrando por id_cia IN (42, 399).
    Si la consulta falla o no devuelve resultados, usa las variables de entorno
    EPAC_URL, EPAC_USERNAME y EPAC_PASSWORD.

    Returns:
        Tupla (url, usuario, contraseña).

    Raises:
        RuntimeError: Si no se pueden obtener las credenciales por ningún medio.
    """
    if _mysql_error is None:
        try:
            cnx = _mysql_connector.connect(**_construir_params_bd())
            cur = cnx.cursor(dictionary=True)
            cur.execute(
                "SELECT url, user, pass "
                "FROM softline_aseguradoras_claves_web "
                "WHERE id_cia IN (42, 399) AND url IS NOT NULL "
                "LIMIT 1"
            )
            fila = cur.fetchone()
            cur.close()
            cnx.close()
            if fila and fila.get("url") and fila.get("user") and fila.get("pass"):
                return str(fila["url"]), str(fila["user"]), str(fila["pass"])
        except Exception as exc:
            print(f"Aviso BD: {exc}. Usando credenciales del entorno.")

    url = os.environ.get("EPAC_URL", "")
    usuario = os.environ.get("EPAC_USERNAME", "")
    contrasena = os.environ.get("EPAC_PASSWORD", "")
    if not url or not usuario or not contrasena:
        raise RuntimeError(
            "No se pudieron obtener las credenciales de ePAC. "
            "Define EPAC_URL, EPAC_USERNAME y EPAC_PASSWORD en el entorno."
        )
    return url, usuario, contrasena


def _navegar_a_en_curso(page: Page) -> None:
    """Hace clic en la pestaña 'En curso' y espera a que cargue el listado.

    Las pestañas son celdas <td>. Al hacer clic en 'En curso' el portal
    carga los resultados vía AJAX. La cabecera de la tabla ya estaba visible
    en la pestaña anterior, así que esperamos al contador 'X de Y' con Y>0
    para confirmar que los datos reales ya han llegado.

    Args:
        page: Página activa de Playwright.
    """
    frame_loc = page.frame_locator(_IFRAME_SELECTOR)

    tab = frame_loc.get_by_role("cell", name="En curso", exact=True)
    expect(tab).to_be_visible(timeout=30_000)
    tab.click()

    # Esperar a que el contador muestre datos reales (Y > 0 en "X de Y")
    frame = page.frame(name="appArea")
    if frame:
        frame.wait_for_function(
            """() => {
                const tds = Array.from(document.querySelectorAll('td'));
                return tds.some(td => {
                    const t = (td.textContent || '').replace(/[\\s\\u00a0]+/g, ' ').trim();
                    const m = t.match(/^(\\d+) de (\\d+)$/);
                    return m && parseInt(m[2], 10) > 0;
                });
            }""",
            timeout=30_000,
        )
    else:
        page.wait_for_timeout(3_000)


def _extraer_filas_pagina(page: Page) -> list[list[str]]:
    """Extrae las filas visibles en la página actual del listado.

    Usa JavaScript puro dentro del iframe para localizar la tabla por su
    cabecera "Stro./Encargo" (puede ser th o td según el portal) y leer
    todas las filas de datos.

    Args:
        page: Página activa de Playwright.

    Returns:
        Lista de listas, cada sublista con los valores de celda de una fila.
    """
    frame = page.frame(name="appArea")
    if not frame:
        return []

    resultado: list[list[str]] = frame.evaluate("""
        () => {
            const cabecera = Array.from(document.querySelectorAll('th, td')).find(
                el => (el.textContent || '').trim() === 'Stro./Encargo'
            );
            if (!cabecera) return [];
            const tabla = cabecera.closest('table');
            if (!tabla) return [];
            const filas = [];
            for (const tr of tabla.querySelectorAll('tr')) {
                const tds = Array.from(tr.querySelectorAll('td'));
                if (tds.length === 0) continue;
                const vals = tds.map(td => (td.innerText || td.textContent || '').trim());
                const texto = vals.join(' ');
                if (!texto.trim() || texto.includes('No hay datos para mostrar')) continue;
                filas.push(vals);
            }
            return filas;
        }
    """) or []
    return resultado


def _leer_contador(page: Page) -> tuple[int, int]:
    """Lee el contador de paginación 'X de Y' del portal.

    Args:
        page: Página activa de Playwright.

    Returns:
        Tupla (mostrado_hasta, total). Devuelve (-1, -1) si no se puede leer.
    """
    frame = page.frame(name="appArea")
    if not frame:
        return -1, -1

    texto: str = frame.evaluate("""
        () => {
            const tds = Array.from(document.querySelectorAll('td'));
            for (const td of tds) {
                const t = (td.textContent || '').replace(/[\\s\\u00a0]+/g, ' ').trim();
                if (/^\\d+ de \\d+$/.test(t)) return t;
            }
            return '';
        }
    """) or ""

    m = re.search(r"(\d+)\s+de\s+(\d+)", texto)
    if m:
        return int(m.group(1)), int(m.group(2))
    return -1, -1


def _avanzar_pagina(page: Page) -> None:
    """Hace clic en 'Avanzar' con el mecanismo nativo de Playwright.

    Usa get_by_role para garantizar que se disparan todos los eventos del portal.

    Args:
        page: Página activa de Playwright.
    """
    page.frame_locator(_IFRAME_SELECTOR).get_by_role(
        "cell", name="Avanzar", exact=True
    ).last.click()


def extraer_todos_registros(page: Page) -> list[list[str]]:
    """Extrae todos los registros del listado paginado navegando por todas las páginas.

    Para detectar el fin usa el contador 'X de Y' del portal. Tras cada clic
    en Avanzar espera un tiempo fijo y comprueba que el contador avanzó.

    Args:
        page: Página activa de Playwright.

    Returns:
        Lista completa de todas las filas extraídas de todas las páginas.
    """
    todos: list[list[str]] = []
    pagina = 1

    while True:
        print(f"  Extrayendo página {pagina}...")
        filas = _extraer_filas_pagina(page)
        todos.extend(filas)

        mostrado, total = _leer_contador(page)
        if pagina == 1:
            print(f"  Total en portal: {total} registros")

        if total <= 0 or mostrado < 0 or mostrado >= total:
            break

        _avanzar_pagina(page)

        # Esperar a que el contador avance (hasta 15 s, chequeando cada 500 ms)
        nuevo_mostrado = mostrado
        deadline = time.monotonic() + 15.0
        while time.monotonic() < deadline:
            page.wait_for_timeout(500)
            nm, _ = _leer_contador(page)
            if nm > mostrado:
                nuevo_mostrado = nm
                break

        print(f"  Contador: {mostrado} → {nuevo_mostrado} (total {total})")
        if nuevo_mostrado <= mostrado:
            print("  Aviso: el contador no avanzó. Fin de paginación.")
            break

        pagina += 1

    print(f"  Total filas extraídas: {len(todos)}")
    return todos


def guardar_excel(filas: list[list[str]], ruta_salida: Path) -> None:
    """Guarda los registros extraídos en un fichero Excel con cabeceras.

    Args:
        filas: Filas extraídas del listado de ePAC.
        ruta_salida: Ruta donde se guardará el fichero Excel.
    """
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "En curso"

    for col_idx, nombre in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.font = Font(bold=True)
        letra = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letra].width = _ANCHOS_COLUMNA.get(nombre, 15)

    for fila_idx, fila in enumerate(filas, start=2):
        for col_idx in range(1, len(COLUMNAS) + 1):
            valor = fila[col_idx - 1] if col_idx - 1 < len(fila) else ""
            ws.cell(row=fila_idx, column=col_idx, value=valor)

    wb.save(ruta_salida)
    print(f"Excel guardado en: {ruta_salida}")


def _parsear_args() -> argparse.Namespace:
    """Define y parsea los argumentos de línea de comandos.

    Returns:
        Namespace con los argumentos parseados.
    """
    p = argparse.ArgumentParser(
        description=(
            "Exporta el listado 'En curso' de Peritaciones Diversos (ePAC) a Excel."
        )
    )
    p.add_argument(
        "--env",
        default=None,
        help="Ruta a fichero .env alternativo.",
    )
    p.add_argument(
        "--headless",
        action="store_true",
        help="Ejecutar en modo sin ventana gráfica.",
    )
    p.add_argument(
        "--out",
        default=None,
        help="Ruta de salida del Excel. Por defecto: <raíz proyecto>/estadisticas_preliminares.xlsx",
    )
    return p.parse_args()


def main() -> None:
    """Punto de entrada principal del script.

    Returns:
        None.
    """
    args = _parsear_args()
    _cargar_env(args.env)

    if args.out:
        ruta_salida = Path(args.out)
    else:
        ruta_salida = OUTPUT_DIR / "estadisticas_preliminares.xlsx"

    print("Obteniendo credenciales de ePAC...")
    url_epac, usuario, contrasena = obtener_credenciales_epac()
    print(f"URL ePAC: {url_epac}")

    config = load_config(
        overrides={
            "base_url": url_epac,
            "username": usuario,
            "password": contrasena,
            "headless": args.headless,
        }
    )

    print("Lanzando navegador...")
    with launch_browser(config) as (_, page):
        print("Iniciando sesión en ePAC...")
        login_pg = LoginPage(page)
        login_pg.open(config.base_url)
        login_pg.login(username=config.username, password=config.password)
        page.wait_for_selector('a[role="menuitem"]', timeout=30_000)

        print("Navegando a Peritaciones Diversos...")
        nav = NavigationPage(page, config)
        nav.goto_informe_pericial_diversos_sea()

        print("Seleccionando pestaña 'En curso'...")
        _navegar_a_en_curso(page)

        print("Extrayendo registros del listado paginado...")
        filas = extraer_todos_registros(page)

    if not filas:
        print("Aviso: no se encontraron registros en el listado 'En curso'.")

    guardar_excel(filas, ruta_salida)
    print("Proceso completado.")


if __name__ == "__main__":
    main()
